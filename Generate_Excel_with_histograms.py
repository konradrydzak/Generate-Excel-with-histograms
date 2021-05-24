import numpy as np
import openpyxl

if __name__ == "__main__":
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, 6 + 1):
        results = {}
        for x in range(1, 1000000 + 1):
            rand = sum([np.random.randint(1, 1000 + 1) for _ in range(i)])
            results.setdefault(rand, 0)
            results[rand] += 1
        for value in results:
            sheet.cell(row=1, column=i * 3 - 2, value="x" + str(i))
            sheet.cell(row=1, column=i * 3 - 1, value="y" + str(i))
            sheet.cell(row=value + 1, column=i * 3 - 2, value=value)
            sheet.cell(row=value + 1, column=i * 3 - 1, value=results[value])
        xvalues = openpyxl.chart.Reference(sheet, min_col=i * 3 - 2, min_row=2, max_row=6 * 1000 + 1)
        yvalues = openpyxl.chart.Reference(sheet, min_col=i * 3 - 1, min_row=2, max_row=6 * 1000 + 1)
        seriesObj = openpyxl.chart.Series(yvalues, xvalues, title=str(i) + ' series')
        chartObj = openpyxl.chart.ScatterChart()
        chartObj.title = 'My ' + str(i) + ' chart'
        chartObj.append(seriesObj)
        sheet.add_chart(chartObj, 'T' + str(15 * (i - 1) + 1))
        print(str(i) + ' part done.')
    wb.save('Histograms.xlsx')
